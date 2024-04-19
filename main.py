import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

from copy import copy
# словарь для удобной работы с колоннами экселя
c = {
    "A": 1,
    "B": 2,
    "C": 3,
    "D": 4,
    "E": 5,
    "F": 6,
    "G": 7,
    "H": 8,
    "I": 9,
    "J": 10,
    "K": 11,
    "L": 12,
    "M": 13,
    "N": 14,
    "O": 15,
    "P": 16,
    "Q": 17,
    "R": 18,
    "S": 19,
    "T": 20,
    "U": 21,
    "V": 22,
    "W": 23,
    "X": 24,
    "Y": 25,
    "Z": 26,
    "AA": 27,
    "AB": 28,
    "AC": 29,
    "AD": 30,
    "AE": 31,
    "AF": 32,
    "AG": 33,
    "AH": 34,
    "AI": 35,
    "AJ": 36,
    "AK": 37,
    "AL": 38,
    "AM": 39,
    "AN": 40,
    "AO": 41,
    "AP": 42,
    "AQ": 43,
    "AR": 44,
    "AS": 45,
    "AT": 46,
    "AU": 47,
    "AV": 48,
    "AW": 49,
    "AX": 50,
    "AY": 51,
}

# Получает список строк, которые нужно найти, и возвращает список их id
# Используется для объеденынх строк 'Поставщики', 'Поставщики - Wisk Telecom Solutions, TOO', 'Покупатели - Wisk Telecom Solutions, TOO'
def get_ids(sheet, arr):
    ids = []
    for i in arr:
        target_row_index = None
        for row_index in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_index, column=1).value == i:
                target_row_index = row_index
                break
        ids.append(target_row_index)
    return ids

# Объединяет ячейки в строке
def merge_cells_in_row(sheet, row_index):
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=21)


# Возвращает номер строки, которую нужно скопировать(пока не знаю какую строку нужно выбирать)
def get_row_to_copy():
    return 25727


# Возвращает номер строки в которую нужно вставить скопированную строку
def get_row_to_paste():
    return 6

def select_files():
    source_file_path = filedialog.askopenfilename(title="Выберите файл для копирования")
    target_file_path = filedialog.askopenfilename(title="Выберите файл для вставки")
    return source_file_path, target_file_path

def select_rows():
    source_row_index = int(input("Введите номер строки для копирования: "))
    target_row_index = int(input("Введите номер строки для вставки: "))
    return source_row_index, target_row_index



def copy_cell_format(source_cell, target_cell):
    """
    Copies the background color and font style from the source cell to the target cell.
    :param source_cell: Source cell to copy format from
    :param target_cell: Target cell to apply the copied format
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.border = copy(source_cell.border)
        target_cell.protection = copy(source_cell.protection)

def copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, source_column, target_column):
    source_cell = source_sheet.cell(row=source_row_index, column=source_column)
    target_cell_style = target_sheet.cell(row=target_row_index + 1, column=target_column)
    target_cell = target_sheet.cell(row=target_row_index, column=target_column)
    target_cell.value = source_cell.value
    copy_cell_format(target_cell_style, target_cell)

def copy_and_paste_rows(source_file_path, target_file_path, source_row_index, target_row_index):
    try:
        source_wb = load_workbook(source_file_path)
        source_sheet = source_wb.active
        target_wb = load_workbook(target_file_path)
        target_sheet = target_wb.active

        ids_to_merge = get_ids(target_sheet, ['Поставщики', 'Поставщики - Wisk Telecom Solutions, TOO',
                                              'Покупатели - Wisk Telecom Solutions, TOO'])

        # Сдвигаем строки ниже на одну, начиная с последней строки
        target_sheet.insert_rows(target_row_index)

        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["A"], c["A"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["B"], c["B"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["C"], c["C"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["D"], c["D"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["E"], c["E"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["F"], c["F"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["G"], c["G"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["H"], c["H"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["I"], c["I"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["J"], c["J"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["K"], c["K"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["L"], c["L"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["M"], c["M"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["N"], c["N"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["O"], c["O"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["P"], c["P"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["Q"], c["Q"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["R"], c["R"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["S"], c["S"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["T"], c["T"])
        copy_cell(source_sheet, target_sheet, source_row_index, target_row_index, c["U"], c["U"])

        target_wb.save(target_file_path)
        # target_wb.save("New.xlsx")
        print("Операция успешно выполнена!")
    except PermissionError:
        print("Ошибка доступа. Убедитесь, что файл для вставки закрыт.")
    except AttributeError:
        print("Произошла ошибка: Значения некоторых ячеек содержат недопустимые символы.")

# strs_to_merge = ['Поставщики', 'Поставщики - Wisk Telecom Solutions, TOO', 'Покупатели - Wisk Telecom Solutions, TOO']
#
# # Открытие таблицы откуда будут копировать
# source_wb = load_workbook('excel_data/7.15.2_1 (43).xlsx')
# source_sheet = source_wb.active
#
# # Открытие таблицы куда будут копировать
# target_wb = load_workbook('excel_data/Реестр_договоров_Аврора_Сервис_и_Виск_телеком_на_31_12_2023_г_.xlsx')
# target_sheet = target_wb.active
#
#
# ids_to_merge = get_ids(target_sheet, strs_to_merge)
#
# source_row_index = get_row_to_copy()
# target_row_index = get_row_to_paste()
#
# target_sheet.insert_rows(target_row_index)
#
#
# target_sheet.cell(row=target_row_index, column=c["A"]).value = source_sheet.cell(row=source_row_index, column=c["A"]).value
# target_sheet.cell(row=target_row_index, column=c["B"]).value = source_sheet.cell(row=source_row_index, column=c["B"]).value
# target_sheet.cell(row=target_row_index, column=c["C"]).value = source_sheet.cell(row=source_row_index, column=c["C"]).value
# target_sheet.cell(row=target_row_index, column=c["D"]).value = source_sheet.cell(row=source_row_index, column=c["D"]).value
# target_sheet.cell(row=target_row_index, column=c["E"]).value = source_sheet.cell(row=source_row_index, column=c["E"]).value
# target_sheet.cell(row=target_row_index, column=c["F"]).value = source_sheet.cell(row=source_row_index, column=c["F"]).value
# target_sheet.cell(row=target_row_index, column=c["G"]).value = source_sheet.cell(row=source_row_index, column=c["G"]).value
# target_sheet.cell(row=target_row_index, column=c["H"]).value = source_sheet.cell(row=source_row_index, column=c["H"]).value
# target_sheet.cell(row=target_row_index, column=c["I"]).value = source_sheet.cell(row=source_row_index, column=c["I"]).value
# target_sheet.cell(row=target_row_index, column=c["J"]).value = source_sheet.cell(row=source_row_index, column=c["J"]).value
# target_sheet.cell(row=target_row_index, column=c["K"]).value = source_sheet.cell(row=source_row_index, column=c["K"]).value
# target_sheet.cell(row=target_row_index, column=c["L"]).value = source_sheet.cell(row=source_row_index, column=c["L"]).value
# target_sheet.cell(row=target_row_index, column=c["M"]).value = source_sheet.cell(row=source_row_index, column=c["M"]).value
# target_sheet.cell(row=target_row_index, column=c["N"]).value = source_sheet.cell(row=source_row_index, column=c["N"]).value
# target_sheet.cell(row=target_row_index, column=c["O"]).value = source_sheet.cell(row=source_row_index, column=c["O"]).value
# target_sheet.cell(row=target_row_index, column=c["P"]).value = source_sheet.cell(row=source_row_index, column=c["P"]).value
# target_sheet.cell(row=target_row_index, column=c["Q"]).value = source_sheet.cell(row=source_row_index, column=c["Q"]).value
# target_sheet.cell(row=target_row_index, column=c["R"]).value = source_sheet.cell(row=source_row_index, column=c["R"]).value
# target_sheet.cell(row=target_row_index, column=c["S"]).value = source_sheet.cell(row=source_row_index, column=c["S"]).value
# target_sheet.cell(row=target_row_index, column=c["T"]).value = source_sheet.cell(row=source_row_index, column=c["T"]).value
# target_sheet.cell(row=target_row_index, column=c["U"]).value = source_sheet.cell(row=source_row_index, column=c["U"]).value

# Объединяем ячейки в целевом листе
# for i in ids_to_merge:
#     if i >= get_row_to_paste():
#         merge_cells_in_row(target_sheet, i+1)
#     else:
#         pass
#
# target_wb.save('excel_data/NEW.xlsx')
#

def main():
    root = tk.Tk()
    root.withdraw()  # Скрыть основное окно Tkinter

    source_file_path, target_file_path = select_files()
    source_row_index, target_row_index = select_rows()

    copy_and_paste_rows(source_file_path, target_file_path, source_row_index, target_row_index)


if __name__ == "__main__":
    main()